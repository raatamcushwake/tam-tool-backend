from fastapi import APIRouter, UploadFile, File, HTTPException
import openpyxl
import re
from io import BytesIO
from datetime import datetime, date

router = APIRouter()

def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ["%d-%b-%y", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except:
                continue
    return None

def extract_latest_date(val):
    if not val:
        return None
    # If already a date/datetime object, return directly
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Otherwise parse string — split by newline or comma for multiple dates
    parts = str(val).replace("\n", ",").split(",")
    dates = []
    for part in parts:
        part = part.strip()
        # Extract just the date token, ignoring trailing labels like "- Vishal"
        match = re.match(r"^(\d{1,4}[/-]\d{1,2}[/-]\d{1,4})", part)
        if match:
            part = match.group(1)
        else:
            part = part.split(" ")[0].split("T")[0]
        d = parse_date(part)
        if d:
            dates.append(d)
    return max(dates) if dates else None

def compute_status(due_date_val, status_val, today):
    status_str = str(status_val or "").strip().lower()
    if status_str in ["closed", "done"]:
        return "Closed", 0
    if status_str in ["on going", "ongoing"]:
        return "On Going", 0
    if status_str == "pending":
        due = extract_latest_date(due_date_val)
        if due and (today - due).days > 0:
            return "Pending", (today - due).days
        return "Pending", 0
    return "NA", 0

@router.post("/parse")
async def parse_cs_tracker(file: UploadFile = File(...)):
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        today = date.today()

        # Read header row to map column names
        headers = {}
        for col_idx, cell in enumerate(ws[1]):
            if cell.value:
                headers[str(cell.value).strip().lower()] = col_idx

        def get_col(row, name):
            idx = headers.get(name.lower())
            if idx is None:
                return ""
            val = row[idx]
            return val if val is not None else ""

        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            sr_no = get_col(row, "s.no")
            description = str(get_col(row, "description") or "").strip()
            if not description:
                continue

            due_date_val = get_col(row, "due date")
            compliance_date_val = get_col(row, "compliance date")
            status_val = get_col(row, "status")
            remarks_val = str(get_col(row, "remarks") or "").strip()

            # Format compliance date — may contain multiple dates (take the latest)
            compliance_date_parsed = extract_latest_date(compliance_date_val)
            if compliance_date_parsed:
                compliance_date_str = compliance_date_parsed.strftime("%d/%m/%Y")
            elif compliance_date_val:
                compliance_date_str = str(compliance_date_val).strip()
            else:
                compliance_date_str = ""

            # Format due date
            due_date_parsed = extract_latest_date(due_date_val)
            if due_date_parsed:
                due_date_str = due_date_parsed.strftime("%d/%m/%Y")
            elif due_date_val:
                raw = str(due_date_val).strip()
                due_date_str = raw.split(" ")[0].split("T")[0]
            else:
                due_date_str = ""

            computed_status, overdue_days = compute_status(due_date_val, status_val, today)

            # Days left for expiring soon
            days_left = 0

            items.append({
                "sr_no": sr_no,
                "description": description,
                "due_date": due_date_str,
                "compliance_date": compliance_date_str,
                "original_status": str(status_val or "").strip(),
                "computed_status": computed_status,
                "overdue_days": overdue_days,
                "days_left": days_left,
                "remarks": remarks_val,
            })

        # Flag pending rows that still need a maker remark before submission
        for item in items:
            item["needs_remark"] = (item["computed_status"] == "Pending")

        summary = {
            "total": len(items),
            "pending": sum(1 for i in items if i["computed_status"] == "Pending"),
            "pending_overdue": sum(1 for i in items if i["computed_status"] == "Pending" and i["overdue_days"] > 0),
            "on_going": sum(1 for i in items if i["computed_status"] == "On Going"),
            "closed": sum(1 for i in items if i["computed_status"] == "Closed"),
            "na": sum(1 for i in items if i["computed_status"] == "NA"),
        }

        return {"status": "success", "items": items, "summary": summary}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))