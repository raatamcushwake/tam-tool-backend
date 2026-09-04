import io
import json
from urllib.parse import quote
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.firebase import get_storage_bucket, get_firestore_client

router = APIRouter()


def resolve_project_name(project_id: str) -> str:
    """Look up the project's display name from Firestore, given its document ID.
    Falls back to the raw project_id if no matching project is found."""
    db = get_firestore_client()
    doc = db.collection("projects").document(project_id).get()
    if doc.exists:
        data = doc.to_dict()
        name = data.get("projectName") or data.get("name")
        if name:
            return name
    return project_id

# ── Structure (must mirror frontend/src/pages/CostReview.jsx SECTIONS) ──
SECTIONS = [
    {
        "title": "Preliminary",
        "items": [
            ("Pre", "BOUNDRY & FANCING"),
            ("Pre", "BORWEL & SOIL TESTING"),
            ("Pre", "Equipments"),
            ("Pre", "Safety"),
        ],
    },
    {
        "title": "RCC and Civil",
        "subsections": [
            {
                "title": "RCC",
                "items": [
                    ("RCC", "Excavation & pilling"),
                    ("RCC", "Steel"),
                    ("RCC", "Concrete"),
                    ("RCC", "Shuttering: Resi"),
                    ("RCC", "Shuttering: Non Resi"),
                    ("RCC", "Labor charges"),
                    ("RCC", "Antitermite"),
                    ("RCC", "Soiling & PCC"),
                ],
            },
            {
                "title": "Civil",
                "items": [
                    ("Civil", "waterproofing"),
                    ("Civil", "Blockwork: Resi"),
                    ("Civil", "Blockwork: Non Resi"),
                    ("Civil", "Internal plaster: Resi"),
                    ("Civil", "Internal plaster: Non Resi"),
                    ("Civil", "External Plaster: Resi"),
                    ("Civil", "External Plaster: Non Resi"),
                ],
            },
        ],
    },
    {
        "title": "Finishing",
        "items": [
            ("FINISHING", "Flooring: Resi"),
            ("FINISHING", "Dado: Resi"),
            ("FINISHING", "Flooring: Non Resi"),
            ("FINISHING", "Kitchen platform"),
            ("FINISHING", "Toilet counter"),
            ("FINISHING", "Window sill"),
            ("FINISHING", "Gypsum/ Patra  punning"),
            ("FINISHING", "Door frame & shutter"),
            ("FINISHING", "Door frame & shutter"),
            ("FINISHING", "Windows"),
            ("FINISHING", "Internal Paint: Resi"),
            ("FINISHING", "Internal Paint: Non Resi"),
            ("FINISHING", "External Paint"),
            ("FINISHING", "S/c railing"),
            ("FINISHING", "Window SS railing"),
            ("FINISHING", "Fabrication: shutters"),
            ("FINISHING", "False-ceiling"),
        ],
    },
    {
        "title": "MEP",
        "items": [
            ("MEP", "Plumbing: conceal"),
            ("MEP", "Plumbing: Fixtures"),
            ("MEP", "Electrical: Conceal"),
            ("MEP", "Electrical: fixtures"),
            ("MEP", "Elevator"),
            ("MEP", "Pumps & Panels"),
            ("MEP", "Stacker / tower parking"),
            ("MEP", "Fire Fighting"),
        ],
    },
    {
        "title": "Infra",
        "items": [
            ("Infra", "STP"),
            ("Infra", "DG SET"),
            ("Infra", "Lighting protection system"),
            ("Infra", "CCTV"),
            ("Infra", "Compound Paving"),
            ("Infra", "Compound wall"),
            ("Infra", "Security cabin"),
            ("Infra", "Gate"),
            ("Infra", "SWD & Service chambers"),
            ("Infra", "Tremix in parking area"),
            ("Infra", "Meter room"),
            ("Infra", "Sinages, Letter boxes, name plates & Building logo"),
            ("Infra", "RWH"),
        ],
    },
    {
        "title": "Amenities",
        "items": [
            ("Amenities", "Gym"),
            ("Amenities", "Changing room - Gents & Ladies"),
            ("Amenities", "BMS"),
            ("Amenities", "Panel Room"),
            ("Amenities", "Landscape"),
            ("Amenities", "Entrace Lobby"),
            ("Amenities", "Façade: louvers"),
            ("Amenities", "Building illumination"),
        ],
    },
    {
        "title": "Misc.",
        "items": [
            ("Misc.", "Electricity"),
            ("Misc.", "Water"),
            ("Misc.", "Security"),
            ("Misc.", "Site Preparation"),
            ("Misc.", "Misc O/H"),
            ("Misc.", "mathadi"),
        ],
    },
]


def flatten_section(section):
    if "subsections" in section:
        items = []
        for sub in section["subsections"]:
            items.extend(sub["items"])
        return items
    return section["items"]


FLAT_ITEMS = [item for section in SECTIONS for item in flatten_section(section)]

STORAGE_PATH_TEMPLATE = "projects/{project_name}/TDD/CostReview.xlsx"

# ── Mirrors PROGRESS_ACTIVITY_MAP / PROGRESS_DESC_MAP in CostReview.jsx —
# maps a Cost Review Work Head/Work Description to the matching Activity
# name from Project Progress, so budgets can be summed per activity. ──────
PROGRESS_ACTIVITY_MAP = {
    "RCC": "RCC",
}

PROGRESS_DESC_MAP = {
    "waterproofing": "Waterproofing",
    "Blockwork: Resi": "Blockwork",
    "Blockwork: Non Resi": "Blockwork",
    "Internal plaster: Resi": "Internal Plaster",
    "Internal plaster: Non Resi": "Internal Plaster",
    "External Plaster: Resi": "Ext Plaster",
    "External Plaster: Non Resi": "Ext Plaster",
    "Flooring: Resi": "Flooring- main",
    "Flooring: Non Resi": "Flooring- main",
    "Gypsum/ Patra  punning": "Gypsum (except staircase)",
    "Door frame & shutter": "Door Shutters",
    "Kitchen platform": "Kitchen Platform",
    "Internal Paint: Resi": "Internal Paint",
    "Internal Paint: Non Resi": "Internal Paint",
    "External Paint": "External Painting",
    "Elevator": "lift Installation",
    "Plumbing: conceal": "Plumbing Downtakes and looping",
    "Fire Fighting": "Fire fighting & FF",
    "Electrical: Conceal": "Electrical Wiring",
    "Electrical: fixtures": "Electrical Fittings",
}


def _row_budget(row_dict):
    """Budget (in cr) = Quantity x Rate / 1,00,00,000, same formula as the frontend."""
    try:
        qty = float(row_dict.get("quantity") or 0)
    except (TypeError, ValueError):
        qty = 0
    try:
        rate = float(row_dict.get("rate") or 0)
    except (TypeError, ValueError):
        rate = 0
    return (qty * rate) / 1e7

# ── Schema (same shape the frontend already sends/expects) ──────────
class AreaEntry(BaseModel):
    value: str = ""
    unit: str = "sft"
    remarks: str = ""

class Areas(BaseModel):
    nonResi: AreaEntry = AreaEntry()
    resi: AreaEntry = AreaEntry()

class CostReviewRow(BaseModel):
    coeff: str = ""
    quantity: str = ""
    unit: str = ""
    rate: str = ""      # CWI Rate
    remarks: str = ""
    cwiCI: str = ""      # NEW — CWI Cost Incurred
    cwiCTC: str = ""     # NEW — CWI Cost to Complete
    devRate: str = ""    # NEW — Developer Rate
    devCI: str = ""       # NEW — Developer Cost Incurred
    devCTC: str = ""      # NEW — Developer Cost to Complete

class ComparisonRow(BaseModel):
    id: str = ""
    item: str = ""
    devRate: str = ""
    devBudget: str = ""
    devCI: str = ""
    devCTC: str = ""
    cwiRate: str = ""
    cwiBudget: str = ""
    cwiCI: str = ""
    cwiCTC: str = ""

class CustomItem(BaseModel):
    id: str = ""
    head: str = ""
    desc: str = ""

class CustomSection(BaseModel):
    id: str = ""
    title: str = ""
    items: List[CustomItem] = []

class CostReviewPayload(BaseModel):
    projectId: str
    areas: Areas
    rows: List[CostReviewRow]
    comparisonRows: List[ComparisonRow] = []
    customSections: List[CustomSection] = []
    excludedAutoRowKeys: List[str] = []
    excludedAutoSectionKeys: List[str] = []
    hiddenSectionKeys: List[str] = []
    hiddenSubKeys: List[str] = []
    hiddenItemKeys: List[str] = []


# ── Workbook builder ─────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")        # table header row
SECTION_FILL = PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid")       # main particular
SUBSECTION_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")    # sub particular
SUBTOTAL_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")      # subsection / flat-section subtotal
SECTION_TOTAL_FILL = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid") # grouped-section Total
GRAND_TOTAL_FILL = PatternFill(start_color="60A5FA", end_color="60A5FA", fill_type="solid")   # grand total
AREA_FILL = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")          # area-summary rows
DEV_HEADER_FILL = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")    # Developer group header (comparison sheet)
CWI_HEADER_FILL = PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid")    # CWI group header (comparison sheet)

BOLD = Font(bold=True)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")


def _write_row(ws, row_num, values):
    """Write a list of values into an explicit row number instead of relying
    on ws.append(), which can drift out of sync with row_cursor and collide
    with a previously merged row (raising 'MergedCell...read-only')."""
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col)
        cell.value = val


def _apply_row_style(ws, row, fill=None, font=None, ncols=14):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        if fill:
            cell.fill = fill
        if font:
            cell.font = font


def build_workbook(
    areas: Areas,
    rows: List[CostReviewRow],
    comparison_rows: List[ComparisonRow] = None,
    custom_sections: List[CustomSection] = None,
    excluded_auto_row_keys: List[str] = None,
    excluded_auto_section_keys: List[str] = None,
    hidden_section_keys: List[str] = None,
    hidden_sub_keys: List[str] = None,
    hidden_item_keys: List[str] = None,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "CWI Cost Working"

    # Area summary (rows 1-3)
    non_resi_val = float(areas.nonResi.value) if areas.nonResi.value not in ("", None) else 0
    resi_val = float(areas.resi.value) if areas.resi.value not in ("", None) else 0
    ws.append(["Non Resi Area", non_resi_val, areas.nonResi.unit or "sft", areas.nonResi.remarks])
    ws.append(["Resi Area", resi_val, areas.resi.unit or "sft", areas.resi.remarks])
    ws.append(["Total", "=B1+B2", "sft", ""])
    for r in range(1, 4):
        ws.cell(row=r, column=1).font = BOLD
        _apply_row_style(ws, r, fill=AREA_FILL, ncols=4)
        ws.row_dimensions[r].height = 22
        ws.cell(row=r, column=1).alignment = WRAP_TOP_LEFT
        ws.cell(row=r, column=4).alignment = WRAP_TOP_LEFT  # Remarks
        ws.cell(row=r, column=2).number_format = "0.00"
    ws.cell(row=3, column=1).fill = SUBTOTAL_FILL
    ws.cell(row=3, column=2).fill = SUBTOTAL_FILL

    ws.append([])  # blank row 4

    # Table header (row 5) — 14 columns now:
    # A Work Head | B Work Description | C Coeff | D Quantity | E Unit
    # F CWI Rate | G CWI Budget | H CWI CI | I CWI CTC
    # J Dev Rate | K Dev Budget | L Dev CI | M Dev CTC | N Remarks
    ws.append([
        "Work Head", "Work Description", "Coeff", "Quantity", "Unit",
        "CWI Rate", "CWI Budget (in cr)", "CWI CI", "CWI CTC",
        "Dev Rate", "Dev Budget (in cr)", "Dev CI", "Dev CTC",
        "Remarks",
    ])
    _apply_row_style(ws, 5, fill=HEADER_FILL, font=BOLD)
    ws.row_dimensions[5].height = 20

    row_cursor = 6
    idx = 0
    cwi_grand_total_components = []
    dev_grand_total_components = []

    def write_item_row(head, desc, r, row_num):
        coeff_val = float(r.coeff) if r.coeff not in ("", None) else ""
        cwi_rate_val = float(r.rate) if r.rate not in ("", None) else ""
        cwi_ci_val = float(r.cwiCI) if r.cwiCI not in ("", None) else ""
        cwi_ctc_val = float(r.cwiCTC) if r.cwiCTC not in ("", None) else ""
        dev_rate_val = float(r.devRate) if r.devRate not in ("", None) else ""
        dev_ci_val = float(r.devCI) if r.devCI not in ("", None) else ""
        dev_ctc_val = float(r.devCTC) if r.devCTC not in ("", None) else ""
        _write_row(ws, row_num, [
            head,
            desc,
            coeff_val,
            float(r.quantity) if r.quantity not in ("", None) else "",
            r.unit,
            cwi_rate_val,
            f'=IF(OR(D{row_num}="",F{row_num}=""),0,D{row_num}*F{row_num}/10000000)',
            cwi_ci_val,
            cwi_ctc_val,
            dev_rate_val,
            f'=IF(OR(D{row_num}="",J{row_num}=""),0,D{row_num}*J{row_num}/10000000)',
            dev_ci_val,
            dev_ctc_val,
            r.remarks,
        ])
        _apply_row_style(ws, row_num)
        ws.cell(row=row_num, column=1).alignment = WRAP_TOP_LEFT   # Work Head
        ws.cell(row=row_num, column=2).alignment = WRAP_TOP_LEFT   # Work Description
        ws.cell(row=row_num, column=14).alignment = WRAP_TOP_LEFT  # Remarks
        ws.row_dimensions[row_num].height = 30
        for col in (3, 4, 6, 7, 8, 9, 10, 11, 12, 13):
            ws.cell(row=row_num, column=col).number_format = "0.00"

    for section in SECTIONS:
        # Section header row
        _write_row(ws, row_cursor, [section["title"]])
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=14)
        hdr_cell = ws.cell(row=row_cursor, column=1)
        hdr_cell.font = Font(bold=True, size=12)
        hdr_cell.alignment = Alignment(vertical="center")
        _apply_row_style(ws, row_cursor, fill=SECTION_FILL)
        ws.row_dimensions[row_cursor].height = 20
        row_cursor += 1

        if "subsections" in section:
            cwi_subtotal_cells = []
            dev_subtotal_cells = []
            for sub in section["subsections"]:
                # Subsection header row
                _write_row(ws, row_cursor, [f"    {sub['title']}"])
                ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=14)
                sub_cell = ws.cell(row=row_cursor, column=1)
                sub_cell.font = Font(bold=True, italic=True)
                sub_cell.alignment = Alignment(vertical="center")
                _apply_row_style(ws, row_cursor, fill=SUBSECTION_FILL)
                ws.row_dimensions[row_cursor].height = 18
                row_cursor += 1

                sub_start = row_cursor
                for head, desc in sub["items"]:
                    r = rows[idx] if idx < len(rows) else CostReviewRow()
                    write_item_row(head, desc, r, row_cursor)
                    row_cursor += 1
                    idx += 1
                sub_end = row_cursor - 1

                subtotal_row = row_cursor
                _write_row(ws, subtotal_row, [
                    "", "", "", "", "", f"Subtotal - {sub['title']}",
                    f"=SUM(G{sub_start}:G{sub_end})", "", "",
                    "", f"=SUM(K{sub_start}:K{sub_end})", "", "", "",
                ])
                _apply_row_style(ws, subtotal_row, fill=SUBTOTAL_FILL, font=BOLD)
                ws.row_dimensions[subtotal_row].height = 18
                ws.cell(row=subtotal_row, column=7).number_format = "0.00"
                ws.cell(row=subtotal_row, column=11).number_format = "0.00"
                cwi_subtotal_cells.append(f"G{subtotal_row}")
                dev_subtotal_cells.append(f"K{subtotal_row}")
                row_cursor += 1

            # Section total = sum of subsection subtotals
            section_total_row = row_cursor
            _write_row(ws, section_total_row, [
                "", "", "", "", "", f"Total - {section['title']}",
                f"={'+'.join(cwi_subtotal_cells)}", "", "",
                "", f"={'+'.join(dev_subtotal_cells)}", "", "", "",
            ])
            _apply_row_style(ws, section_total_row, fill=SECTION_TOTAL_FILL, font=Font(bold=True, size=11))
            ws.row_dimensions[section_total_row].height = 20
            ws.cell(row=section_total_row, column=7).number_format = "0.00"
            ws.cell(row=section_total_row, column=11).number_format = "0.00"
            cwi_grand_total_components.append(f"G{section_total_row}")
            dev_grand_total_components.append(f"K{section_total_row}")
            row_cursor += 1

        else:
            section_start = row_cursor
            for head, desc in section["items"]:
                r = rows[idx] if idx < len(rows) else CostReviewRow()
                write_item_row(head, desc, r, row_cursor)
                row_cursor += 1
                idx += 1
            section_end = row_cursor - 1

            subtotal_row = row_cursor
            _write_row(ws, subtotal_row, [
                "", "", "", "", "", f"Subtotal - {section['title']}",
                f"=SUM(G{section_start}:G{section_end})", "", "",
                "", f"=SUM(K{section_start}:K{section_end})", "", "", "",
            ])
            _apply_row_style(ws, subtotal_row, fill=SUBTOTAL_FILL, font=BOLD)
            ws.row_dimensions[subtotal_row].height = 18
            ws.cell(row=subtotal_row, column=7).number_format = "0.00"
            ws.cell(row=subtotal_row, column=11).number_format = "0.00"
            cwi_grand_total_components.append(f"G{subtotal_row}")
            dev_grand_total_components.append(f"K{subtotal_row}")
            row_cursor += 1

    # ── Custom sections (Maker-added Work Heads, always appended at the end) ──
    for section in (custom_sections or []):
        _write_row(ws, row_cursor, [section.title])
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=14)
        hdr_cell = ws.cell(row=row_cursor, column=1)
        hdr_cell.font = Font(bold=True, size=12)
        hdr_cell.alignment = Alignment(vertical="center")
        _apply_row_style(ws, row_cursor, fill=SECTION_FILL)
        ws.row_dimensions[row_cursor].height = 20
        row_cursor += 1

        section_start = row_cursor
        for item in section.items:
            r = rows[idx] if idx < len(rows) else CostReviewRow()
            write_item_row(item.head, item.desc, r, row_cursor)
            row_cursor += 1
            idx += 1
        section_end = row_cursor - 1

        if section_end >= section_start:
            subtotal_row = row_cursor
            _write_row(ws, subtotal_row, [
                "", "", "", "", "", f"Subtotal - {section.title}",
                f"=SUM(G{section_start}:G{section_end})", "", "",
                "", f"=SUM(K{section_start}:K{section_end})", "", "", "",
            ])
            _apply_row_style(ws, subtotal_row, fill=SUBTOTAL_FILL, font=BOLD)
            ws.row_dimensions[subtotal_row].height = 18
            ws.cell(row=subtotal_row, column=7).number_format = "0.00"
            ws.cell(row=subtotal_row, column=11).number_format = "0.00"
            cwi_grand_total_components.append(f"G{subtotal_row}")
            dev_grand_total_components.append(f"K{subtotal_row}")
            row_cursor += 1

    grand_row = row_cursor
    _write_row(ws, grand_row, [
        "", "", "", "", "", "Grand Total",
        f"={'+'.join(cwi_grand_total_components)}", "", "",
        "", f"={'+'.join(dev_grand_total_components)}", "", "", "",
    ])
    _apply_row_style(ws, grand_row, fill=GRAND_TOTAL_FILL, font=Font(bold=True, size=12, color="FFFFFF"))
    ws.row_dimensions[grand_row].height = 22
    ws.cell(row=grand_row, column=7).number_format = "0.00"
    ws.cell(row=grand_row, column=11).number_format = "0.00"

    widths = {
        "A": 14, "B": 42, "C": 10, "D": 14, "E": 10,
        "F": 12, "G": 16, "H": 12, "I": 12,
        "J": 12, "K": 16, "L": 12, "M": 12, "N": 32,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"

    build_comparison_sheet(wb, comparison_rows or [])
    build_config_sheet(
        wb,
        excluded_auto_row_keys or [],
        excluded_auto_section_keys or [],
        hidden_section_keys or [],
        hidden_sub_keys or [],
        hidden_item_keys or [],
    )

    return wb


def build_config_sheet(
    wb: Workbook,
    excluded_row_keys: List[str],
    excluded_section_keys: List[str],
    hidden_section_keys: List[str] = None,
    hidden_sub_keys: List[str] = None,
    hidden_item_keys: List[str] = None,
):
    """Hidden sheet that stores which rows/Work Heads the Maker has 'cut'
    from the Redirect-from-Cost-Working comparison view, plus which fixed
    sections/subsections/items are soft-hidden in CWI Cost Working itself,
    so both survive Save → reload."""
    ws = wb.create_sheet("_config")
    ws.sheet_state = "hidden"
    ws["A1"] = "excludedAutoRowKeys"
    ws["B1"] = json.dumps(excluded_row_keys)
    ws["A2"] = "excludedAutoSectionKeys"
    ws["B2"] = json.dumps(excluded_section_keys)
    ws["A3"] = "hiddenSectionKeys"
    ws["B3"] = json.dumps(hidden_section_keys or [])
    ws["A4"] = "hiddenSubKeys"
    ws["B4"] = json.dumps(hidden_sub_keys or [])
    ws["A5"] = "hiddenItemKeys"
    ws["B5"] = json.dumps(hidden_item_keys or [])


def build_comparison_sheet(wb: Workbook, comparison_rows: List[ComparisonRow]):
    """Second sheet: free-form, manually entered — Item | Developer (Rate, Budget, CI, CTC) | CWI (Rate, Budget, CI, CTC)."""
    ws = wb.create_sheet("Comparison Report")

    ws.merge_cells("A1:I1")
    ws["A1"] = "Comparison Report — Developer vs CWI"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A3:A4")
    ws["A3"] = "Item"
    ws.merge_cells("B3:E3")
    ws["B3"] = "Developer"
    ws.merge_cells("F3:I3")
    ws["F3"] = "CWI"
    for col, fill in ((1, HEADER_FILL), (2, DEV_HEADER_FILL), (6, CWI_HEADER_FILL)):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill
    _apply_row_style(ws, 3, ncols=9)
    ws.row_dimensions[3].height = 20

    sub_headers = ["Rate", "Budget (in cr)", "CI", "CTC", "Rate", "Budget (in cr)", "CI", "CTC"]
    for c, label in enumerate(sub_headers, start=2):  # start at column 2 — column 1 (A4) is
                                                        # merged into A3 ("Item") and read-only
        ws.cell(row=4, column=c, value=label)
    _apply_row_style(ws, 4, fill=HEADER_FILL, font=BOLD, ncols=9)
    ws.row_dimensions[4].height = 20

    def num_or_blank(v):
        try:
            return float(v) if v not in ("", None) else ""
        except (TypeError, ValueError):
            return ""

    row_cursor = 5
    for r in comparison_rows:
        _write_row(ws, row_cursor, [
            r.item,
            num_or_blank(r.devRate), num_or_blank(r.devBudget), num_or_blank(r.devCI), num_or_blank(r.devCTC),
            num_or_blank(r.cwiRate), num_or_blank(r.cwiBudget), num_or_blank(r.cwiCI), num_or_blank(r.cwiCTC),
        ])
        _apply_row_style(ws, row_cursor, ncols=9)
        ws.cell(row=row_cursor, column=1).alignment = WRAP_TOP_LEFT
        ws.row_dimensions[row_cursor].height = 24
        for col in (2, 3, 4, 5, 6, 7, 8, 9):
            ws.cell(row=row_cursor, column=col).number_format = "0.00"
        row_cursor += 1

    first_data_row, last_data_row = 5, row_cursor - 1
    grand_row = row_cursor
    dev_formula = f"=SUM(C{first_data_row}:C{last_data_row})" if last_data_row >= first_data_row else 0
    cwi_formula = f"=SUM(G{first_data_row}:G{last_data_row})" if last_data_row >= first_data_row else 0

    _write_row(ws, grand_row, ["Grand Total", "", dev_formula, "", "", "", cwi_formula, "", ""])
    _apply_row_style(ws, grand_row, fill=GRAND_TOTAL_FILL, font=Font(bold=True, size=12, color="FFFFFF"), ncols=9)
    ws.row_dimensions[grand_row].height = 22
    ws.cell(row=grand_row, column=3).number_format = "0.00"
    ws.cell(row=grand_row, column=7).number_format = "0.00"

    widths = {"A": 42, "B": 12, "C": 16, "D": 10, "E": 10, "F": 12, "G": 16, "H": 10, "I": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

def parse_workbook(wb: Workbook) -> dict:
    ws = wb["CWI Cost Working"] if "CWI Cost Working" in wb.sheetnames else wb.active

    def cell_str(row, col):
        v = ws.cell(row=row, column=col).value
        return "" if v is None else str(v)

    areas = {
        "nonResi": {"value": cell_str(1, 2), "unit": cell_str(1, 3) or "sft", "remarks": cell_str(1, 4)},
        "resi": {"value": cell_str(2, 2), "unit": cell_str(2, 3) or "sft", "remarks": cell_str(2, 4)},
    }

    rows = [None] * len(FLAT_ITEMS)
    occurrences_by_key = {}
    for i, key in enumerate(FLAT_ITEMS):
        occurrences_by_key.setdefault(key, []).append(i)
    seen_counts = {}

    custom_rows = []
    custom_sections_out = []
    current_custom_section = None

    def _cell_row_dict(row):
        coeff = row[2].value
        qty = row[3].value
        unit = row[4].value
        rate = row[5].value
        cwi_ci = row[7].value
        cwi_ctc = row[8].value
        dev_rate = row[9].value
        dev_ci = row[11].value
        dev_ctc = row[12].value
        remarks = row[13].value
        return {
            "coeff": "" if coeff is None else str(coeff),
            "quantity": "" if qty is None else str(qty),
            "unit": "" if unit is None else str(unit),
            "rate": "" if rate is None else str(rate),
            "remarks": "" if remarks is None else str(remarks),
            "cwiCI": "" if cwi_ci is None else str(cwi_ci),
            "cwiCTC": "" if cwi_ctc is None else str(cwi_ctc),
            "devRate": "" if dev_rate is None else str(dev_rate),
            "devCI": "" if dev_ci is None else str(dev_ci),
            "devCTC": "" if dev_ctc is None else str(dev_ctc),
        }

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        head_val = row[0].value
        desc_val = row[1].value
        if not head_val or not desc_val:
            continue
        key = (str(head_val), str(desc_val))
        occurrences = occurrences_by_key.get(key)

        if occurrences:
            count = seen_counts.get(key, 0)
            if count >= len(occurrences):
                continue
            idx = occurrences[count]
            seen_counts[key] = count + 1
            rows[idx] = _cell_row_dict(row)
            continue

        # Not a known fixed item -> a Maker-added custom row, always after
        # the fixed items in sheet order, so appending preserves position.
        head_str = str(head_val)
        if current_custom_section is None or current_custom_section["title"] != head_str:
            current_custom_section = {
                "id": f"cs_{len(custom_sections_out)}",
                "title": head_str,
                "items": [],
            }
            custom_sections_out.append(current_custom_section)

        current_custom_section["items"].append({
            "id": f"ci_{len(custom_rows)}",
            "head": head_str,
            "desc": str(desc_val),
        })
        custom_rows.append(_cell_row_dict(row))

    empty_row = {"coeff": "", "quantity": "", "unit": "", "rate": "", "remarks": "",
                  "cwiCI": "", "cwiCTC": "", "devRate": "", "devCI": "", "devCTC": ""}
    rows = [r if r is not None else dict(empty_row) for r in rows]
    rows = rows + custom_rows

    comparison_rows = []
    if "Comparison Report" in wb.sheetnames:
        cws = wb["Comparison Report"]
        for row in cws.iter_rows(min_row=5, max_row=cws.max_row):
            item_val = row[0].value
            if not item_val or str(item_val).strip().lower() == "grand total":
                continue
            comparison_rows.append({
                "id": "",
                "item": str(item_val),
                "devRate": "" if row[1].value is None else str(row[1].value),
                "devBudget": "" if row[2].value is None else str(row[2].value),
                "devCI": "" if row[3].value is None else str(row[3].value),
                "devCTC": "" if row[4].value is None else str(row[4].value),
                "cwiRate": "" if row[5].value is None else str(row[5].value),
                "cwiBudget": "" if row[6].value is None else str(row[6].value),
                "cwiCI": "" if row[7].value is None else str(row[7].value),
                "cwiCTC": "" if row[8].value is None else str(row[8].value),
            })

    excluded_auto_row_keys = []
    excluded_auto_section_keys = []
    hidden_section_keys = []
    hidden_sub_keys = []
    hidden_item_keys = []
    if "_config" in wb.sheetnames:
        cfg = wb["_config"]
        try:
            excluded_auto_row_keys = json.loads(cfg["B1"].value) if cfg["B1"].value else []
        except (TypeError, ValueError):
            excluded_auto_row_keys = []
        try:
            excluded_auto_section_keys = json.loads(cfg["B2"].value) if cfg["B2"].value else []
        except (TypeError, ValueError):
            excluded_auto_section_keys = []
        try:
            hidden_section_keys = json.loads(cfg["B3"].value) if cfg["B3"].value else []
        except (TypeError, ValueError):
            hidden_section_keys = []
        try:
            hidden_sub_keys = json.loads(cfg["B4"].value) if cfg["B4"].value else []
        except (TypeError, ValueError):
            hidden_sub_keys = []
        try:
            hidden_item_keys = json.loads(cfg["B5"].value) if cfg["B5"].value else []
        except (TypeError, ValueError):
            hidden_item_keys = []

    return {
        "areas": areas,
        "rows": rows,
        "comparisonRows": comparison_rows,
        "customSections": custom_sections_out,
        "excludedAutoRowKeys": excluded_auto_row_keys,
        "excludedAutoSectionKeys": excluded_auto_section_keys,
        "hiddenSectionKeys": hidden_section_keys,
        "hiddenSubKeys": hidden_sub_keys,
        "hiddenItemKeys": hidden_item_keys,
    }


# ── Routes ────────────────────────────────────────────────────────────
@router.get("/{project_id}")
async def get_cost_review(project_id: str):
    try:
        project_name = resolve_project_name(project_id)
        bucket = get_storage_bucket()
        blob = bucket.blob(STORAGE_PATH_TEMPLATE.format(project_name=project_name))
        if not blob.exists():
            return {
                "projectId": project_id,
                "areas": {},
                "rows": [],
                "comparisonRows": [],
                "excludedAutoRowKeys": [],
                "excludedAutoSectionKeys": [],
                "hiddenSectionKeys": [],
                "hiddenSubKeys": [],
                "hiddenItemKeys": [],
                "status": "DRAFT",
            }
        data = blob.download_as_bytes()
        wb = load_workbook(io.BytesIO(data), data_only=False)
        parsed = parse_workbook(wb)
        return {"projectId": project_id, **parsed, "status": "SAVED"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{project_id}")
async def save_cost_review(project_id: str, payload: CostReviewPayload):
    print(">>> USING CostReviewService.py save_cost_review <<<")
    try:
        project_name = resolve_project_name(project_id)
        wb = build_workbook(
            payload.areas,
            payload.rows,
            payload.comparisonRows,
            payload.customSections,
            payload.excludedAutoRowKeys,
            payload.excludedAutoSectionKeys,
            payload.hiddenSectionKeys,
            payload.hiddenSubKeys,
            payload.hiddenItemKeys,
        )
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        bucket = get_storage_bucket()
        blob = bucket.blob(STORAGE_PATH_TEMPLATE.format(project_name=project_name))
        blob.upload_from_file(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return {"message": "Cost Review saved to Storage", "path": blob.name}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{project_id}/export")
async def export_cost_review(project_id: str, payload: CostReviewPayload):
    """Builds the colorful, wrap-text .xlsx directly from whatever is currently
    on screen (no need to Save first) and streams it back as a real Excel file.
    Now includes both the 'CWI Cost Working' sheet and the 'Comparison Report' sheet."""
    try:
        project_name = resolve_project_name(project_id)
        wb = build_workbook(
            payload.areas,
            payload.rows,
            payload.comparisonRows,
            payload.customSections,
            payload.excludedAutoRowKeys,
            payload.excludedAutoSectionKeys,
            payload.hiddenSectionKeys,
            payload.hiddenSubKeys,
            payload.hiddenItemKeys,
        )
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"CostReview_{project_name}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
                "Access-Control-Expose-Headers": "Content-Disposition",
            },
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{project_id}/activity-budgets")
async def get_activity_budgets(project_id: str):
    """Sums each fixed row's CWI Budget by matching Activity name (same
    mapping used for % Progress), plus the overall grand total. Project
    Progress uses this to auto-calculate each activity's Cost Weightage %
    as (activity budget / grand total) x 100."""
    try:
        project_name = resolve_project_name(project_id)
        bucket = get_storage_bucket()
        blob = bucket.blob(STORAGE_PATH_TEMPLATE.format(project_name=project_name))
        if not blob.exists():
            return {"budgets": {}, "grandTotal": 0}

        data = blob.download_as_bytes()
        wb = load_workbook(io.BytesIO(data), data_only=False)
        parsed = parse_workbook(wb)
        rows = parsed["rows"]

        budgets = {}
        grand_total = 0.0
        for i, (head, desc) in enumerate(FLAT_ITEMS):
            row = rows[i] if i < len(rows) else {}
            budget = _row_budget(row)
            grand_total += budget
            activity = PROGRESS_DESC_MAP.get(desc) or PROGRESS_ACTIVITY_MAP.get(head)
            if activity:
                budgets[activity] = budgets.get(activity, 0) + budget

        return {
            "budgets": {k: round(v, 4) for k, v in budgets.items()},
            "grandTotal": round(grand_total, 4),
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{project_id}/download-url")
async def get_download_url(project_id: str):
    """Optional: lets the frontend link straight to the last-Saved workbook
    in Storage instead of re-generating it on the fly via /export."""
    try:
        project_name = resolve_project_name(project_id)
        bucket = get_storage_bucket()
        blob = bucket.blob(STORAGE_PATH_TEMPLATE.format(project_name=project_name))
        if not blob.exists():
            raise HTTPException(status_code=404, detail="No Cost Review saved yet")
        url = blob.generate_signed_url(expiration=3600, method="GET")
        return {"url": url}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
