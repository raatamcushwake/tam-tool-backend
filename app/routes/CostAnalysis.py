import io
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.firebase import get_storage_bucket, get_firestore_client

router = APIRouter()


def resolve_project_name(project_id: str) -> str:
    """Look up the project's display name from Firestore, given its document ID.
    Falls back to the raw project_id if no matching project is found."""
    db = get_firestore_client()
    doc = db.collection("projects").document(project_id).get()
    if doc.exists:
        data = doc.to_dict()
        name = data.get("projectName") or data.get("name")
        if name:
            return name
    return project_id

# ── Structure (must mirror frontend/src/pages/CostReview.jsx SECTIONS) ──
SECTIONS = [
    {
        "title": "RCC, Civil & Preliminary",
        "subsections": [
            {
                "title": "Preliminary",
                "items": [
                    ("Pre", "BOUNDRY & FANCING"),
                    ("Pre", "BORWEL & SOIL TESTING"),
                    ("Pre", "Equipments"),
                    ("Pre", "Safety"),
                ],
            },
            {
                "title": "RCC",
                "items": [
                    ("RCC", "Excavation & pilling"),
                    ("RCC", "Steel"),
                    ("RCC", "Concrete"),
                    ("RCC", "Shuttering: Resi"),
                    ("RCC", "Shuttering: Non Resi"),
                    ("RCC", "Labor charges"),
                    ("RCC", "Antitermite"),
                    ("RCC", "Soiling & PCC"),
                ],
            },
            {
                "title": "Civil",
                "items": [
                    ("Civil", "waterproofing"),
                    ("Civil", "Blockwork: Resi"),
                    ("Civil", "Blockwork: Non Resi"),
                    ("Civil", "Internal plaster: Resi"),
                    ("Civil", "Internal plaster: Non Resi"),
                    ("Civil", "External Plaster: Resi"),
                    ("Civil", "External Plaster: Non Resi"),
                ],
            },
        ],
    },
    {
        "title": "Finishing",
        "items": [
            ("FINISHING", "Flooring: Resi"),
            ("FINISHING", "Dado: Resi"),
            ("FINISHING", "Flooring: Non Resi"),
            ("FINISHING", "Kitchen platform"),
            ("FINISHING", "Toilet counter"),
            ("FINISHING", "Window sill"),
            ("FINISHING", "Gypsum/ Patra  punning"),
            ("FINISHING", "Door frame & shutter"),
            ("FINISHING", "Door frame & shutter"),
            ("FINISHING", "Windows"),
            ("FINISHING", "Internal Paint: Resi"),
            ("FINISHING", "Internal Paint: Non Resi"),
            ("FINISHING", "External Paint"),
            ("FINISHING", "S/c railing"),
            ("FINISHING", "Window SS railing"),
            ("FINISHING", "Fabrication: shutters"),
            ("FINISHING", "False-ceiling"),
        ],
    },
    {
        "title": "MEP",
        "items": [
            ("MEP", "Plumbing: conceal"),
            ("MEP", "Plumbing: Fixtures"),
            ("MEP", "Electrical: Conceal"),
            ("MEP", "Electrical: fixtures"),
            ("MEP", "Elevator"),
            ("MEP", "Pumps & Panels"),
            ("MEP", "Stacker / tower parking"),
            ("MEP", "Fire Fighting"),
        ],
    },
    {
        "title": "Infra",
        "items": [
            ("Infra", "STP"),
            ("Infra", "DG SET"),
            ("Infra", "Lighting protection system"),
            ("Infra", "CCTV"),
            ("Infra", "Compound Paving"),
            ("Infra", "Compound wall"),
            ("Infra", "Security cabin"),
            ("Infra", "Gate"),
            ("Infra", "SWD & Service chambers"),
            ("Infra", "Tremix in parking area"),
            ("Infra", "Meter room"),
            ("Infra", "Sinages, Letter boxes, name plates & Building logo"),
            ("Infra", "RWH"),
        ],
    },
    {
        "title": "Amenities",
        "items": [
            ("Amenities", "Gym"),
            ("Amenities", "Changing room - Gents & Ladies"),
            ("Amenities", "BMS"),
            ("Amenities", "Panel Room"),
            ("Amenities", "Landscape"),
            ("Amenities", "Entrace Lobby"),
            ("Amenities", "Façade: louvers"),
            ("Amenities", "Building illumination"),
        ],
    },
    {
        "title": "Misc.",
        "items": [
            ("Misc.", "Electricity"),
            ("Misc.", "Water"),
            ("Misc.", "Security"),
            ("Misc.", "Site Preparation"),
            ("Misc.", "Misc O/H"),
            ("Misc.", "mathadi"),
        ],
    },
]


def flatten_section(section):
    if "subsections" in section:
        items = []
        for sub in section["subsections"]:
            items.extend(sub["items"])
        return items
    return section["items"]


FLAT_ITEMS = [item for section in SECTIONS for item in flatten_section(section)]

STORAGE_PATH_TEMPLATE = "projects/{project_name}/TDD/CostReview.xlsx"

# ── Schema (same shape the frontend already sends/expects) ──────────
class AreaEntry(BaseModel):
    value: str = ""
    unit: str = "sft"
    remarks: str = ""

class Areas(BaseModel):
    nonResi: AreaEntry = AreaEntry()
    resi: AreaEntry = AreaEntry()

class CostReviewRow(BaseModel):
    coeff: str = ""
    quantity: str = ""
    unit: str = ""
    rate: str = ""
    remarks: str = ""

class CostReviewPayload(BaseModel):
    projectId: str
    areas: Areas
    rows: List[CostReviewRow]


# ── Workbook builder ─────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")       # table header row
SECTION_FILL = PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid")      # main particular
SUBSECTION_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")   # sub particular
SUBTOTAL_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")     # subsection / flat-section subtotal
SECTION_TOTAL_FILL = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")# grouped-section Total
GRAND_TOTAL_FILL = PatternFill(start_color="60A5FA", end_color="60A5FA", fill_type="solid")  # grand total

BOLD = Font(bold=True)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def _apply_row_style(ws, row, fill=None, font=None, ncols=8):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        if fill:
            cell.fill = fill
        if font:
            cell.font = font


def build_workbook(areas: Areas, rows: List[CostReviewRow]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Review"

    # Area summary (rows 1-3)
    ws.append(["Non Resi Area", areas.nonResi.value or 0, areas.nonResi.unit or "sft", areas.nonResi.remarks])
    ws.append(["Resi Area", areas.resi.value or 0, areas.resi.unit or "sft", areas.resi.remarks])
    ws.append(["Total", "=B1+B2", "sft", ""])
    for r in range(1, 4):
        ws.cell(row=r, column=1).font = BOLD
        _apply_row_style(ws, r, ncols=4)
    ws.cell(row=3, column=1).fill = SUBTOTAL_FILL
    ws.cell(row=3, column=2).fill = SUBTOTAL_FILL

    ws.append([])  # blank row 4

    # Table header (row 5)
    ws.append(["Work Head", "Work Description", "Coeff", "Quantity", "Unit", "Rate", "Budget (in cr)", "Remarks"])
    _apply_row_style(ws, 5, fill=HEADER_FILL, font=BOLD)

    row_cursor = 6
    idx = 0
    grand_total_components = []

    def write_item_row(head, desc, r, row_num):
        ws.append([
            head,
            desc,
            r.coeff,
            float(r.quantity) if r.quantity not in ("", None) else "",
            r.unit,
            float(r.rate) if r.rate not in ("", None) else "",
            f'=IF(OR(D{row_num}="",F{row_num}=""),0,D{row_num}*F{row_num}/10000000)',
            r.remarks,
        ])
        _apply_row_style(ws, row_num)
        ws.cell(row=row_num, column=2).alignment = WRAP_TOP  # Work Description, fully visible
        ws.cell(row=row_num, column=8).alignment = WRAP_TOP  # Remarks, fully visible

    for section in SECTIONS:
        # Section header row
        ws.append([section["title"]])
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)
        hdr_cell = ws.cell(row=row_cursor, column=1)
        hdr_cell.font = Font(bold=True, size=12)
        _apply_row_style(ws, row_cursor, fill=SECTION_FILL)
        row_cursor += 1

        if "subsections" in section:
            subtotal_cells = []
            for sub in section["subsections"]:
                # Subsection header row
                ws.append([f"    {sub['title']}"])
                ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)
                sub_cell = ws.cell(row=row_cursor, column=1)
                sub_cell.font = Font(bold=True, italic=True)
                _apply_row_style(ws, row_cursor, fill=SUBSECTION_FILL)
                row_cursor += 1

                sub_start = row_cursor
                for head, desc in sub["items"]:
                    r = rows[idx] if idx < len(rows) else CostReviewRow()
                    write_item_row(head, desc, r, row_cursor)
                    row_cursor += 1
                    idx += 1
                sub_end = row_cursor - 1

                subtotal_row = row_cursor
                ws.append(["", "", "", "", "", f"Subtotal — {sub['title']}", f"=SUM(G{sub_start}:G{sub_end})", ""])
                _apply_row_style(ws, subtotal_row, fill=SUBTOTAL_FILL, font=BOLD)
                subtotal_cells.append(f"G{subtotal_row}")
                row_cursor += 1

            # Section total = sum of subsection subtotals
            section_total_row = row_cursor
            ws.append(["", "", "", "", "", f"Total — {section['title']}", f"={'+'.join(subtotal_cells)}", ""])
            _apply_row_style(ws, section_total_row, fill=SECTION_TOTAL_FILL, font=Font(bold=True, size=11))
            grand_total_components.append(f"G{section_total_row}")
            row_cursor += 1

        else:
            section_start = row_cursor
            for head, desc in section["items"]:
                r = rows[idx] if idx < len(rows) else CostReviewRow()
                write_item_row(head, desc, r, row_cursor)
                row_cursor += 1
                idx += 1
            section_end = row_cursor - 1

            subtotal_row = row_cursor
            ws.append(["", "", "", "", "", f"Subtotal — {section['title']}", f"=SUM(G{section_start}:G{section_end})", ""])
            _apply_row_style(ws, subtotal_row, fill=SUBTOTAL_FILL, font=BOLD)
            grand_total_components.append(f"G{subtotal_row}")
            row_cursor += 1

    grand_row = row_cursor
    ws.append(["", "", "", "", "", "Grand Total", f"={'+'.join(grand_total_components)}", ""])
    _apply_row_style(ws, grand_row, fill=GRAND_TOTAL_FILL, font=Font(bold=True, size=12, color="FFFFFF"))

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 32
    ws.freeze_panes = "A6"

    return wb


def parse_workbook(wb: Workbook) -> dict:
    ws = wb.active

    def cell_str(row, col):
        v = ws.cell(row=row, column=col).value
        return "" if v is None else str(v)

    areas = {
        "nonResi": {"value": cell_str(1, 2), "unit": cell_str(1, 3) or "sft", "remarks": cell_str(1, 4)},
        "resi": {"value": cell_str(2, 2), "unit": cell_str(2, 3) or "sft", "remarks": cell_str(2, 4)},
    }

    rows = [None] * len(FLAT_ITEMS)
    occurrences_by_key = {}
    for i, key in enumerate(FLAT_ITEMS):
        occurrences_by_key.setdefault(key, []).append(i)
    seen_counts = {}

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        head_val = row[0].value
        desc_val = row[1].value
        # Section / subsection headers are merged (desc cell is None) and
        # subtotal/total rows leave column B empty — both get skipped here.
        if not head_val or not desc_val:
            continue
        key = (str(head_val), str(desc_val))
        occurrences = occurrences_by_key.get(key)
        if not occurrences:
            continue
        count = seen_counts.get(key, 0)
        if count >= len(occurrences):
            continue
        idx = occurrences[count]
        seen_counts[key] = count + 1

        coeff = row[2].value
        qty = row[3].value
        unit = row[4].value
        rate = row[5].value
        remarks = row[7].value
        rows[idx] = {
            "coeff": "" if coeff is None else str(coeff),
            "quantity": "" if qty is None else str(qty),
            "unit": "" if unit is None else str(unit),
            "rate": "" if rate is None else str(rate),
            "remarks": "" if remarks is None else str(remarks),
        }

    rows = [r if r is not None else {"coeff": "", "quantity": "", "unit": "", "rate": "", "remarks": ""} for r in rows]
    return {"areas": areas, "rows": rows}


# ── Routes ────────────────────────────────────────────────────────────
@router.get("/{project_id}")
async def get_cost_review(project_id: str):
    try:
        project_name = resolve_project_name(project_id)
        bucket = get_storage_bucket()
        blob = bucket.blob(STORAGE_PATH_TEMPLATE.format(project_name=project_name))
        if not blob.exists():
            return {"projectId": project_id, "areas": {}, "rows": [], "status": "DRAFT"}
        data = blob.download_as_bytes()
        wb = load_workbook(io.BytesIO(data), data_only=False)
        parsed = parse_workbook(wb)
        return {"projectId": project_id, **parsed, "status": "SAVED"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{project_id}")
async def save_cost_review(project_id: str, payload: CostReviewPayload):
    try:
        project_name = resolve_project_name(project_id)
        wb = build_workbook(payload.areas, payload.rows)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        bucket = get_storage_bucket()
        blob = bucket.blob(STORAGE_PATH_TEMPLATE.format(project_name=project_name))
        blob.upload_from_file(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return {"message": "Cost Review saved to Storage", "path": blob.name}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{project_id}/download-url")
async def get_download_url(project_id: str):
    """Optional: lets the frontend's 'Download Report' button link straight to the
    stored workbook instead of generating a CSV client-side."""
    try:
        project_name = resolve_project_name(project_id)
        bucket = get_storage_bucket()
        blob = bucket.blob(STORAGE_PATH_TEMPLATE.format(project_name=project_name))
        if not blob.exists():
            raise HTTPException(status_code=404, detail="No Cost Review saved yet")
        url = blob.generate_signed_url(expiration=3600, method="GET")
        return {"url": url}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
