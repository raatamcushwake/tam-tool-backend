from fastapi import APIRouter, HTTPException
from app.core.firebase import get_firestore
from pydantic import BaseModel
from typing import Optional
from firebase_admin import firestore
from app.utils.code_generator import generate_service_code
import io
from urllib.parse import quote
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from app.core.firebase import get_storage_bucket

router = APIRouter()

PROGRESS_STORAGE_PATH_TEMPLATE = "projects/{project_name}/TDD/ProjectProgress.xlsx"


def resolve_project_name(project_id: str, db) -> str:
    """Look up the project's display name from Firestore, given its document ID.
    Falls back to the raw project_id if no matching project is found."""
    doc = db.collection("projects").document(project_id).get()
    if doc.exists:
        data = doc.to_dict()
        name = data.get("projectName") or data.get("name")
        if name:
            return name
    return project_id

_BOLD = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
_SUMMARY_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_row(ws, row, ncols, fill=None, font=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = _BORDER
        if fill:
            cell.fill = fill
        if font:
            cell.font = font


def _flatten_activities(packages):
    acts = []
    for p in packages:
        for a in p.get("activities", []):
            acts.append(a)
    return acts


def _activity_weightage_total(packages, activity_name):
    for p in packages:
        for a in p.get("activities", []):
            if a.get("name") == activity_name:
                values = a.get("values", {}) or {}
                return sum(float(v) if v not in ("", None) else 0 for v in values.values())
    return 0


def _activity_cost_weightage(packages, activity_name):
    for p in packages:
        for a in p.get("activities", []):
            if a.get("name") == activity_name:
                return float(a.get("costWeightage") or 0)
    return 0


def _total_entered(matrix_for_tower, activity_name):
    values = (matrix_for_tower or {}).get("values", {}).get(activity_name, {}) or {}
    return sum(float(v) if v not in ("", None) else 0 for v in values.values())


def _activity_completion_percent_for_tower(packages, activity_matrix, tower_name, activity_name):
    denom = _activity_weightage_total(packages, activity_name)
    if not denom:
        return 0
    entered = _total_entered(activity_matrix.get(tower_name), activity_name)
    return (entered / denom) * 100


def _activity_weighted_total_percent(towers, packages, activity_matrix, activity_name):
    """Same number as the 'Total' row at the bottom of Tower Level Status —
    each tower's completion % for this activity, weighted by that tower's
    construction area."""
    total_area = sum(float(t.get("constructionArea") or 0) for t in towers) or 0
    if not total_area:
        return 0
    weighted_sum = 0
    for t in towers:
        name = t.get("name")
        completion = _activity_completion_percent_for_tower(packages, activity_matrix, name, activity_name)
        area = float(t.get("constructionArea") or 0)
        weighted_sum += completion * area
    return weighted_sum / total_area


def build_project_progress_workbook(tower_config: dict, weightage_config: dict, activity_matrix: dict) -> Workbook:
    towers = (tower_config or {}).get("towers", []) or []
    packages = (weightage_config or {}).get("packages", []) or []
    activities = _flatten_activities(packages)
    activity_matrix = activity_matrix or {}

    total_area = sum(float(t.get("constructionArea") or 0) for t in towers) or 0

    def tower_weightage_percent(t):
        if not total_area:
            return 0
        return (float(t.get("constructionArea") or 0) / total_area) * 100

    def activity_completion_percent(tower_name, activity_name):
        denom = _activity_weightage_total(packages, activity_name)
        if not denom:
            return 0
        entered = _total_entered(activity_matrix.get(tower_name), activity_name)
        return (entered / denom) * 100

    def tower_progress_percent(tower_name):
        total = 0
        for a in activities:
            name = a.get("name")
            completion = activity_completion_percent(tower_name, name)
            weightage = _activity_cost_weightage(packages, name)
            total += (completion / 100) * weightage
        return total

    def project_progress_percent():
        total = 0
        for t in towers:
            tw = tower_weightage_percent(t) / 100
            tp = tower_progress_percent(t.get("name"))
            total += tw * tp
        return total

    wb = Workbook()

    ws = wb.active
    ws.title = "Tower Level Status"
    headers = ["Tower"] + [a.get("name") for a in activities] + ["Tower Total"]
    ws.append(headers)
    _style_row(ws, 1, len(headers), fill=_HEADER_FILL, font=_BOLD)

    for t in towers:
        name = t.get("name")
        row = [name]
        for a in activities:
            row.append(round(activity_completion_percent(name, a.get("name")), 2))
        row.append(round(tower_progress_percent(name), 2))
        ws.append(row)
        r = ws.max_row
        _style_row(ws, r, len(headers))
        ws.cell(row=r, column=len(headers)).fill = _SUMMARY_FILL
        ws.cell(row=r, column=len(headers)).font = _BOLD

    ws.column_dimensions["A"].width = 20
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 16

    ws2 = wb.create_sheet("Project Level Summary")
    ws2.append(["Tower", "Weightage %", "Tower Progress %"])
    _style_row(ws2, 1, 3, fill=_HEADER_FILL, font=_BOLD)

    for t in towers:
        name = t.get("name")
        ws2.append([
            name,
            round(tower_weightage_percent(t), 2),
            round(tower_progress_percent(name), 2),
        ])
        _style_row(ws2, ws2.max_row, 3)

    ws2.append([])
    ws2.append(["Project Progress", "", round(project_progress_percent(), 2)])
    _style_row(ws2, ws2.max_row, 3, fill=_SUMMARY_FILL, font=_BOLD)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 18

    return wb


class SimpleProjectRequest(BaseModel):
    name: str
    description: str = ""
    region: str = ""
    enabledModules: list[str] = []
    enabledServices: dict[str, list[str]] = {}
    basicInfo: dict = {}
    creatorUid: str = ""
    creatorName: str = ""
    creatorRole: str = ""  # "ADMIN" or "MANAGER"

@router.post("")
async def create_project_simple(request: SimpleProjectRequest):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document()

        flat_modules = sorted({m for mods in request.enabledServices.values() for m in mods})

        project_data = {
            "id": project_ref.id,
            "name": request.name,
            "projectName": request.name,
            "description": request.description,
            "region": request.region,
            "status": "ACTIVE",
            "members": {},
            "enabledServices": request.enabledServices,
            "enabledModules": flat_modules or request.enabledModules,
            "basicInfo": request.basicInfo,
            "createdBy": request.creatorUid,
            "createdByName": request.creatorName,
            "createdByRole": request.creatorRole,
            "createdAt": firestore.SERVER_TIMESTAMP
        }
        project_ref.set(project_data)

        # A Manager who creates a project is auto-assigned MANAGER on it,
        # one entry per enabled service, so it shows in their own project list.
        if request.creatorUid and request.creatorRole == "MANAGER":
            user_ref = db.collection("users").document(request.creatorUid)
            service_entries = [
                {
                    "projectId": project_ref.id,
                    "projectName": request.name,
                    "role": "MANAGER",
                    "serviceKey": key,
                    "serviceLabel": "",
                }
                for key in request.enabledServices.keys()
            ] or [{
                "projectId": project_ref.id,
                "projectName": request.name,
                "role": "MANAGER",
                "serviceKey": "",
                "serviceLabel": "",
            }]
            user_ref.update({
                "projectRoles": firestore.ArrayUnion(service_entries)
            })

        return {"message": "Project created", "id": project_ref.id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
class ProjectRequest(BaseModel):
    name: str
    description: str = ""
    makerId: str
    reviewerId: str
    managerId: str
    enabledModules: list[str] = []

class AssignRoleRequest(BaseModel):
    userId: str
    projectId: str
    role: str  # MAKER, REVIEWER, MANAGER

@router.post("/create")
async def create_project(request: ProjectRequest):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document()
        project_data = {
            "id": project_ref.id,
            "name": request.name,
            "projectName": request.name,
            "description": request.description,
            "status": "ACTIVE",
            "members": {
                "makerId": request.makerId,
                "reviewerId": request.reviewerId,
                "managerId": request.managerId
            },
            "enabledModules": request.enabledModules,
            "createdAt": firestore.SERVER_TIMESTAMP
        }
        project_ref.set(project_data)

        # Assign roles to users
        for uid, role in [
            (request.makerId, "MAKER"),
            (request.reviewerId, "REVIEWER"),
            (request.managerId, "MANAGER")
        ]:
            user_ref = db.collection("users").document(uid)
            user_ref.update({
                "projectRoles": firestore.ArrayUnion([{
                    "projectId": project_ref.id,
                    "role": role
                }])
            })

        return {"message": "Project created", "projectId": project_ref.id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("")
async def get_all_projects():
    try:
        db = get_firestore()
        docs = list(db.collection("projects").stream())
        projects = []
        for d in docs:
            data = d.to_dict()
            region = data.get("region") or "WES"
            enabled_services = data.get("enabledServices", {})
            codes = data.get("codes", {})

            updated = False
            for service_key in enabled_services:
                if service_key not in codes:
                    code = generate_service_code(db, region, service_key, data.get("createdAt"))
                    if code:
                        codes[service_key] = code
                        updated = True

            if updated:
                db.collection("projects").document(d.id).update({"codes": codes})
                data["codes"] = codes

            projects.append(data)
        return projects
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/user/{uid}")
async def get_user_projects(uid: str):
    try:
        db = get_firestore()
        user_doc = db.collection("users").document(uid).get()
        if not user_doc.exists:
            raise HTTPException(status_code=404, detail="User not found")
        user_data = user_doc.to_dict()
        project_roles = user_data.get("projectRoles", [])
        projects = []
        for pr in project_roles:
            proj = db.collection("projects").document(pr["projectId"]).get()
            if proj.exists:
                data = proj.to_dict()
                data["myRole"] = pr["role"]
                projects.append(data)
        return projects
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/{project_id}")
async def get_project(project_id: str):
    try:
        db = get_firestore()
        doc = db.collection("projects").document(project_id).get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")
        return doc.to_dict()
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

class UpdateBasicInfoRequest(BaseModel):
    name: str = None
    region: str = None
    basicInfo: dict = {}

@router.patch("/{project_id}/basic-info")
async def update_project_basic_info(project_id: str, request: UpdateBasicInfoRequest):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        if not project_ref.get().exists:
            raise HTTPException(status_code=404, detail="Project not found")

        update_data = {"basicInfo": request.basicInfo}
        if request.name:
            update_data["name"] = request.name
            update_data["projectName"] = request.name
        if request.region:
            update_data["region"] = request.region

        project_ref.update(update_data)
        return project_ref.get().to_dict()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

class CloseServiceRequest(BaseModel):
    serviceKey: str

@router.patch("/{project_id}/close-service")
async def close_project_service(project_id: str, request: CloseServiceRequest):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        if not project_ref.get().exists:
            raise HTTPException(status_code=404, detail="Project not found")

        project_ref.update({
            "closedServices": firestore.ArrayUnion([request.serviceKey])
        })
        return {"message": "Service closed", "serviceKey": request.serviceKey}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.patch("/{project_id}/reopen-service")
async def reopen_project_service(project_id: str, request: CloseServiceRequest):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        if not project_ref.get().exists:
            raise HTTPException(status_code=404, detail="Project not found")

        project_ref.update({
            "closedServices": firestore.ArrayRemove([request.serviceKey])
        })
        return {"message": "Service reopened", "serviceKey": request.serviceKey}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

class UpdateModulesRequest(BaseModel):
    enabledModules: list[str] = []
    enabledServices: dict[str, list[str]] = {}

@router.patch("/{project_id}/modules")
async def update_project_modules(project_id: str, request: UpdateModulesRequest):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        if not project_ref.get().exists:
            raise HTTPException(status_code=404, detail="Project not found")

        flat_modules = sorted({m for mods in request.enabledServices.values() for m in mods})

        project_ref.update({
            "enabledServices": request.enabledServices,
            "enabledModules": flat_modules or request.enabledModules,
        })
        return {"enabledServices": request.enabledServices, "enabledModules": flat_modules}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

from pydantic import field_validator

def _blank_to_zero(v):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return 0
    return v

class TowerConfig(BaseModel):
    name: str
    constructionArea: Optional[float] = 0
    basements: Optional[float] = 0
    ground: Optional[float] = 0
    stilt: Optional[float] = 0
    podiums: Optional[float] = 0
    serviceFloor: Optional[float] = 0
    upperFloors: Optional[float] = 0

    @field_validator(
        "constructionArea", "basements", "ground", "stilt",
        "podiums", "serviceFloor", "upperFloors", mode="before"
    )
    @classmethod
    def _empty_str_to_zero(cls, v):
        return _blank_to_zero(v)

class TowerConfigRequest(BaseModel):
    totalTowerArea: Optional[float] = 0
    nonTowerArea: Optional[float] = 0
    towers: list[TowerConfig] = []

    @field_validator("totalTowerArea", "nonTowerArea", mode="before")
    @classmethod
    def _empty_str_to_zero(cls, v):
        return _blank_to_zero(v)


@router.get("/{project_id}/tower-config")
async def get_tower_config(project_id: str):
    try:
        db = get_firestore()
        doc = db.collection("projects").document(project_id).get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")
        data = doc.to_dict()
        tower_config = data.get("towerConfig")
        if not tower_config:
            raise HTTPException(status_code=404, detail="Tower config not set yet")
        return tower_config
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{project_id}/tower-config")
async def save_tower_config(project_id: str, request: TowerConfigRequest):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        doc = project_ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")

        tower_config = {
            "totalTowerArea": request.totalTowerArea,
            "nonTowerArea": request.nonTowerArea,
            "towers": [t.dict() for t in request.towers],
            "status": "locked",
        }
        project_ref.update({"towerConfig": tower_config})
        return tower_config
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/{project_id}/tower-config/unlock")
async def unlock_tower_config(project_id: str):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        doc = project_ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")
        project_ref.update({"towerConfig.status": "draft"})
        return {"status": "draft"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class NonTowerConfigRequest(BaseModel):
    name: str = "Non Tower Area"
    constructionArea: Optional[float] = 0
    basements: Optional[float] = 0
    ground: Optional[float] = 0
    stilt: Optional[float] = 0
    podiums: Optional[float] = 0
    serviceFloor: Optional[float] = 0
    upperFloors: Optional[float] = 0

    @field_validator(
        "constructionArea", "basements", "ground", "stilt",
        "podiums", "serviceFloor", "upperFloors", mode="before"
    )
    @classmethod
    def _empty_str_to_zero(cls, v):
        return _blank_to_zero(v)


@router.get("/{project_id}/non-tower-config")
async def get_non_tower_config(project_id: str):
    try:
        db = get_firestore()
        doc = db.collection("projects").document(project_id).get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")
        data = doc.to_dict()
        non_tower_config = data.get("nonTowerConfig")
        if not non_tower_config:
            raise HTTPException(status_code=404, detail="Non tower config not set yet")
        return non_tower_config
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{project_id}/non-tower-config")
async def save_non_tower_config(project_id: str, request: NonTowerConfigRequest):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        doc = project_ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")

        non_tower_config = {
            "data": request.dict(),
            "status": "locked",
        }
        project_ref.update({"nonTowerConfig": non_tower_config})
        return non_tower_config
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/{project_id}/non-tower-config/unlock")
async def unlock_non_tower_config(project_id: str):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        doc = project_ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")
        project_ref.update({"nonTowerConfig.status": "draft"})
        return {"status": "draft"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class WeightageActivity(BaseModel):
    name: str
    values: dict = {}
    costWeightage: Optional[float] = 0
    remarks: Optional[str] = ""

class WeightagePackage(BaseModel):
    name: str
    activities: list[WeightageActivity] = []

class WeightageConfigRequest(BaseModel):
    packages: list[WeightagePackage] = []


@router.get("/{project_id}/weightage-config")
async def get_weightage_config(project_id: str):
    try:
        db = get_firestore()
        doc = db.collection("projects").document(project_id).get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")
        config = doc.to_dict().get("weightageConfig")
        if not config:
            raise HTTPException(status_code=404, detail="Weightage config not set yet")
        return config
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{project_id}/weightage-config")
async def save_weightage_config(project_id: str, request: WeightageConfigRequest):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        if not project_ref.get().exists:
            raise HTTPException(status_code=404, detail="Project not found")
        config = {
            "packages": [p.dict() for p in request.packages],
            "status": "locked",
        }
        project_ref.update({"weightageConfig": config})
        return config
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/{project_id}/weightage-config/unlock")
async def unlock_weightage_config(project_id: str):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        if not project_ref.get().exists:
            raise HTTPException(status_code=404, detail="Project not found")
        project_ref.update({"weightageConfig.status": "draft"})
        return {"status": "draft"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

class ActivityMatrixStructureRequest(BaseModel):
    floors: list[dict] = []
    applicability: dict = {}

class ActivityMatrixValuesRequest(BaseModel):
    values: dict = {}


@router.get("/{project_id}/activity-matrix/{tower_name}")
async def get_activity_matrix(project_id: str, tower_name: str):
    try:
        db = get_firestore()
        doc = db.collection("projects").document(project_id).get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")
        matrix = doc.to_dict().get("activityMatrix", {}).get(tower_name)
        if not matrix:
            raise HTTPException(status_code=404, detail="Activity matrix not set yet")
        return matrix
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{project_id}/activity-matrix/{tower_name}")
async def save_activity_matrix_structure(project_id: str, tower_name: str, request: ActivityMatrixStructureRequest):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        doc = project_ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")

        existing = doc.to_dict().get("activityMatrix", {}).get(tower_name, {})
        update_data = {
            f"activityMatrix.{tower_name}.floors": request.floors,
            f"activityMatrix.{tower_name}.applicability": request.applicability,
            f"activityMatrix.{tower_name}.status": "locked",
        }
        # First-ever save for this tower — make sure `values` exists so the frontend never reads undefined
        if "values" not in existing:
            update_data[f"activityMatrix.{tower_name}.values"] = {}

        project_ref.update(update_data)
        doc2 = project_ref.get().to_dict()
        return doc2.get("activityMatrix", {}).get(tower_name)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/{project_id}/activity-matrix/{tower_name}/values")
async def save_activity_matrix_values(project_id: str, tower_name: str, request: ActivityMatrixValuesRequest):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        if not project_ref.get().exists:
            raise HTTPException(status_code=404, detail="Project not found")
        project_ref.update({f"activityMatrix.{tower_name}.values": request.values})
        return {"values": request.values}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/{project_id}/activity-matrix/{tower_name}/unlock")
async def unlock_activity_matrix(project_id: str, tower_name: str):
    try:
        db = get_firestore()
        project_ref = db.collection("projects").document(project_id)
        if not project_ref.get().exists:
            raise HTTPException(status_code=404, detail="Project not found")
        project_ref.update({f"activityMatrix.{tower_name}.status": "draft"})
        return {"status": "draft"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{project_id}/activity-summary")
async def get_activity_summary(project_id: str):
    """Each Activity's Total Completion % — weighted across all towers by
    construction area (the 'Total' row of Tower Level Status). Cost Review
    uses this to show % Progress against matching Work Heads."""
    try:
        db = get_firestore()
        doc = db.collection("projects").document(project_id).get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")
        data = doc.to_dict()

        towers = (data.get("towerConfig") or {}).get("towers", []) or []
        packages = (data.get("weightageConfig") or {}).get("packages", []) or []
        activity_matrix = data.get("activityMatrix") or {}

        summary = {}
        for a in _flatten_activities(packages):
            name = a.get("name")
            summary[name] = round(
                _activity_weighted_total_percent(towers, packages, activity_matrix, name), 1
            )
        return summary
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{project_id}/export")
async def export_project_progress(project_id: str):
    """Builds the Project Progress workbook (Tower Level Status + Project Level
    Summary) from whatever is currently saved in Firestore and uploads it to
    Storage at projects/{project_name}/TDD/ProjectProgress.xlsx, right next to
    CostReview.xlsx."""
    try:
        db = get_firestore()
        doc = db.collection("projects").document(project_id).get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Project not found")
        data = doc.to_dict()

        tower_config = data.get("towerConfig") or {}
        weightage_config = data.get("weightageConfig") or {}
        activity_matrix = data.get("activityMatrix") or {}

        wb = build_project_progress_workbook(tower_config, weightage_config, activity_matrix)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        project_name = resolve_project_name(project_id, db)
        bucket = get_storage_bucket()
        blob = bucket.blob(PROGRESS_STORAGE_PATH_TEMPLATE.format(project_name=project_name))
        blob.upload_from_file(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return {"message": "Project Progress saved to Storage", "path": blob.name}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
