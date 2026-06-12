from fastapi import APIRouter, UploadFile, File, HTTPException, Form
import openpyxl
from io import BytesIO
from datetime import datetime, date

router = APIRouter()

@router.post("/parse-required")
async def parse_required_approvals(file: UploadFile = File(...)):
    """Parse manager's required approvals list Excel"""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            items.append({
                "sr_no": row[0],
                "approval_name": str(row[1] or "").strip(),
                "stage": str(row[2] or "").strip(),
            })
        return { "status": "success", "items": items, "total": len(items) }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/parse-filled")
async def parse_filled_approvals(file: UploadFile = File(...)):
    """Parse maker's filled approvals Excel and compute status"""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        today = date.today()
        items = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            approval_date = row[5]
            validity = str(row[6] or "").strip()
            expiry_date = None
            status = "NA"

            # Parse expiry
            try:
                if approval_date and validity:
                    if isinstance(approval_date, datetime):
                        approval_date = approval_date.date()
                    years = 0
                    if "year" in validity.lower():
                        years = int(''.join(filter(str.isdigit, validity)))
                    elif "month" in validity.lower():
                        months = int(''.join(filter(str.isdigit, validity)))
                        years = months / 12
                    from dateutil.relativedelta import relativedelta
                    expiry_date = approval_date + relativedelta(years=int(years))

                    days_left = (expiry_date - today).days
                    if days_left < 0:
                        status = "Expired"
                    elif days_left <= 90:
                        status = "Expiring Soon"
                    else:
                        status = "Valid"
            except:
                status = "NA"

            items.append({
                "sr_no": row[0],
                "approval_name": str(row[1] or "").strip(),
                "authority_name": str(row[2] or "").strip(),
                "approval_document": str(row[3] or "").strip(),
                "coverage": str(row[4] or "").strip(),
                "approval_date": str(approval_date) if approval_date else "",
                "validity_period": validity,
                "expiry_date": str(expiry_date) if expiry_date else "",
                "status": status,
                "key_observation": str(row[7] or "").strip(),
                "provided_by_developer": str(row[8] or "").strip(),
                "cwi_comments": str(row[9] or "").strip(),
            })

        summary = {
            "total": len(items),
            "valid": sum(1 for i in items if i["status"] == "Valid"),
            "expired": sum(1 for i in items if i["status"] == "Expired"),
            "expiring_soon": sum(1 for i in items if i["status"] == "Expiring Soon"),
            "na": sum(1 for i in items if i["status"] == "NA"),
        }

        return { "status": "success", "items": items, "summary": summary }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@router.post("/compare")
async def compare_approvals(
    required_file: UploadFile = File(...),
    filled_file: UploadFile = File(...)
):
    try:
        req_content = await required_file.read()
        req_wb = openpyxl.load_workbook(BytesIO(req_content))
        req_ws = req_wb.active
        required_names = set()
        for row in req_ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                required_names.add(str(row[1]).strip().lower())

        filled_content = await filled_file.read()
        filled_wb = openpyxl.load_workbook(BytesIO(filled_content))
        filled_ws = filled_wb.active
        filled_names = set()
        for row in filled_ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                filled_names.add(str(row[1]).strip().lower())

        missing = list(required_names - filled_names)
        return {
            "status": "success",
            "missing_approvals": missing,
            "total_missing": len(missing)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))